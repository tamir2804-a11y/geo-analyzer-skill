#!/usr/bin/env python3
"""Generate a professional GEO Analysis Excel file from JSON data."""

import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


def create_geo_excel(data, output_path):
    """
    Create a GEO analysis Excel file.

    data format:
    {
        "sector": "ביטוח",
        "date": "2026-04-13",
        "weights": {
            "chatgpt": 25, "gemini": 20, "google_ai_overview": 25,
            "claude": 10, "grok": 10, "google_organic": 10
        },
        "companies": [
            {
                "name": "Company Name",
                "url": "https://example.com",
                "scores": {
                    "chatgpt": 7, "gemini": 6, "google_ai_overview": 8,
                    "claude": 5, "grok": 4, "google_organic": 7
                },
                "weighted_score": 6.7,
                "failures": "כשל 1; כשל 2",
                "recommendations": "1. המלצה ראשונה\n2. המלצה שנייה"
            }
        ]
    }
    """
    wb = Workbook()

    # --- Sheet 1: GEO Analysis ---
    ws = wb.active
    ws.title = "GEO Analysis"
    ws.sheet_view.rightToLeft = True

    # Colors
    header_fill = PatternFill("solid", fgColor="2B3A67")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Arial", bold=True, size=14, color="2B3A67")
    subtitle_font = Font(name="Arial", size=10, color="666666")
    data_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    sector = data.get("sector", "")
    date = data.get("date", datetime.now().strftime("%Y-%m-%d"))

    # Title row
    ws.merge_cells("A1:L1")
    ws["A1"] = f"ניתוח GEO — סקטור: {sector} | תאריך: {date}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle
    ws.merge_cells("A2:L2")
    ws["A2"] = "Generative Engine Optimization — נוכחות חברות במנועי AI גנרטיביים"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers (row 3)
    headers = ["#", "שם חברה", "לינק לאתר", "ChatGPT", "Gemini",
               "Google AI Overview", "Claude", "Grok", "Google Organic",
               "ציון משוכלל", "Key GEO Failures", "Recommendations"]
    widths = [5, 25, 35, 14, 14, 20, 14, 14, 18, 16, 45, 55]

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 30

    # Sort companies by weighted_score descending
    companies = sorted(data.get("companies", []), key=lambda c: c.get("weighted_score", 0), reverse=True)

    # Data rows
    for i, company in enumerate(companies):
        row = i + 4
        is_alt = i % 2 == 1
        row_fill = alt_fill if is_alt else PatternFill("solid", fgColor="FFFFFF")

        # #
        cell = ws.cell(row=row, column=1, value=i + 1)
        cell.font = data_font
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.border = thin_border

        # Company name
        cell = ws.cell(row=row, column=2, value=company.get("name", ""))
        cell.font = bold_font
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="right", vertical="top")
        cell.border = thin_border

        # URL as hyperlink
        url = company.get("url", "")
        cell = ws.cell(row=row, column=3, value=url)
        if url.startswith("http"):
            cell.hyperlink = url
        cell.font = link_font
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left", vertical="top")
        cell.border = thin_border

        # Scores
        scores = company.get("scores", {})
        score_keys = ["chatgpt", "gemini", "google_ai_overview", "claude", "grok", "google_organic"]
        for j, key in enumerate(score_keys):
            score = scores.get(key, 0)
            cell = ws.cell(row=row, column=4 + j, value=score)
            cell.font = data_font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = thin_border
            cell.number_format = "0.0"

        # Weighted score
        ws_score = company.get("weighted_score", 0)
        cell = ws.cell(row=row, column=10, value=round(ws_score, 1))
        cell.font = Font(name="Arial", bold=True, size=11)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.border = thin_border
        cell.number_format = "0.0"

        # Failures (English — LTR)
        cell = ws.cell(row=row, column=11, value=company.get("failures", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border

        # Recommendations (English — LTR)
        cell = ws.cell(row=row, column=12, value=company.get("recommendations", ""))
        cell.font = data_font
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border

        ws.row_dimensions[row].height = 60

    # Conditional formatting for score columns (D-J)
    last_row = len(companies) + 3
    for col in ["D", "E", "F", "G", "H", "I", "J"]:
        rng = f"{col}4:{col}{last_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["8"], fill=green_fill))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["5", "7.9"], fill=yellow_fill))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["5"], fill=red_fill))

    # Freeze panes and auto-filter
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:L{last_row}"

    # --- Sheet 2: Weights & Methodology ---
    ws2 = wb.create_sheet("Weights & Methodology")
    ws2.sheet_view.rightToLeft = True

    ws2["A1"] = "מתודולוגיה ומשקלות"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:C1")

    ws2["A3"] = "סקטור:"
    ws2["B3"] = sector
    ws2["A4"] = "תאריך ניתוח:"
    ws2["B4"] = date
    ws2["A5"] = "מספר חברות:"
    ws2["B5"] = len(companies)

    for r in range(3, 6):
        ws2.cell(row=r, column=1).font = bold_font
        ws2.cell(row=r, column=2).font = data_font

    # Weights table
    ws2["A7"] = "משקלות הציון המשוכלל"
    ws2["A7"].font = Font(name="Arial", bold=True, size=12, color="2B3A67")

    weight_headers = ["מנוע AI", "משקל (%)"]
    for j, h in enumerate(weight_headers, 1):
        cell = ws2.cell(row=8, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    weights = data.get("weights", {})
    engine_names = {
        "chatgpt": "ChatGPT", "gemini": "Gemini",
        "google_ai_overview": "Google AI Overview",
        "claude": "Claude", "grok": "Grok",
        "google_organic": "Google Organic"
    }
    for i, (key, label) in enumerate(engine_names.items()):
        ws2.cell(row=9 + i, column=1, value=label).font = data_font
        w_cell = ws2.cell(row=9 + i, column=2, value=weights.get(key, 0))
        w_cell.font = data_font
        w_cell.number_format = "0"
        w_cell.alignment = Alignment(horizontal="center")

    # Methodology note
    note_row = 16
    ws2.cell(row=note_row, column=1, value="הערות מתודולוגיות").font = Font(name="Arial", bold=True, size=12, color="2B3A67")
    ws2.merge_cells(f"A{note_row+1}:D{note_row+5}")
    note_cell = ws2.cell(row=note_row + 1, column=1)
    note_cell.value = (
        "הניתוח מבוסס על מחקר אינטרנטי וסימנים עקיפים לנוכחות במנועי AI. "
        "הציונים משקפים הערכה של נוכחות החברה במקורות שמנועי AI שואבים מהם "
        "(כגון Wikipedia, ביקורות, מאמרים, structured data באתר, ועוד). "
        "הניתוח אינו מבוסס על שאילתה ישירה לכל מנוע AI, "
        "ולכן יש לקחת את הציונים כאינדיקציה כללית ולא כמדד מדויק.\n\n"
        "סולם ציונים:\n"
        "9-10: נוכחות בולטת — החברה מומלצת באופן עקבי\n"
        "7-8: נוכחות טובה — מוזכרת בעקביות\n"
        "5-6: נוכחות בינונית — מופיעה לפעמים\n"
        "3-4: נוכחות חלשה — כמעט לא מוזכרת\n"
        "1-2: אין נוכחות — לא מופיעה כלל"
    )
    note_cell.font = Font(name="Arial", size=10)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="right")

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 40

    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_geo_excel.py <json_data_file> <output_xlsx_path>")
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)

    create_geo_excel(data, sys.argv[2])
